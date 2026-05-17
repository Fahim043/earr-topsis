"""
run_real_dataset_benchmarks.py
================================
Benchmark M1-M7 on the real datasets in ./real_datasets.

The script runs two evaluation modes for each converted dataset:
  1. CLEAN: real data converted into the project TOPSIS JSON schema.
  2. CONTAMINATED: the same real data with a synthetic coordinated DM attack.

Important interpretation note:
Some real files are not native group-DM fuzzy TOPSIS panels. For crisp datasets,
the script creates a deterministic pseudo-panel of DMs by adding small bounded
rating noise around the real normalized criterion values. This is suitable for
robustness experiments, but the paper must state this conversion protocol.
"""

from __future__ import annotations

import argparse
import copy
import csv
import json
import math
import random
import warnings
from collections import Counter
from pathlib import Path
from statistics import mean, pstdev

from openpyxl import load_workbook

import m1_normal_topsis as m1
import m2_bagged_topsis as m2
import m3_ml_filtered as m3
import m4_weighted_bagging as m4
import m5_cluster_stratified as m5
import m6_reliability_weighted as m6
import m7_entropy_reliability as m7


warnings.filterwarnings("ignore", category=RuntimeWarning)

ROOT = Path(__file__).resolve().parent
REAL_DIR = ROOT / "real_datasets"
OUT_DIR = ROOT / "outputs" / "real_dataset_benchmarks"

METHODS = [
    ("M1", lambda path, num_bags, seed: m1.run_m1(path), False),
    ("M2", lambda path, num_bags, seed: m2.run_m2(path, num_bags=num_bags, seed=seed + 2), True),
    ("M3", lambda path, num_bags, seed: m3.run_m3(path)[0], False),
    ("M4", lambda path, num_bags, seed: m4.run_m4(path, num_bags=num_bags, seed=seed + 4)[0], True),
    ("M5", lambda path, num_bags, seed: m5.run_m5(path, num_bags=num_bags, seed=seed + 5)[0], True),
    ("M6", lambda path, num_bags, seed: m6.run_m6(path, num_bags=num_bags, seed=seed + 6)[0], True),
    ("M7", lambda path, num_bags, seed: m7.run_m7(path, num_bags=num_bags, seed=seed + 7)[0], True),
]


def tfn_from_score(score: float, spread: float = 0.7) -> list[float]:
    """Convert a 1-9 crisp score into a bounded triangular fuzzy number."""
    score = max(1.0, min(9.0, score))
    return [
        round(max(1.0, score - spread), 4),
        round(score, 4),
        round(min(9.0, score + spread), 4),
    ]


def parse_s_value(value) -> list[float] | None:
    """Parse strings like [s4] or [s3,s4] into a simple TFN."""
    if not isinstance(value, str) or not value.startswith("[s"):
        return None
    nums = []
    for part in value.strip("[]").split(","):
        part = part.strip()
        if part.startswith("s"):
            nums.append(float(part[1:]))
    if not nums:
        return None
    lo, hi = min(nums), max(nums)
    mid = sum(nums) / len(nums)
    # Map 0-6 linguistic scale to 1-9 numeric scale.
    scale = lambda x: 1.0 + (8.0 * x / 6.0)
    return [round(scale(lo), 4), round(scale(mid), 4), round(scale(hi), 4)]


def save_dataset(data: dict, name: str) -> Path:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUT_DIR / f"{name}.json"
    path.write_text(json.dumps(data, indent=2), encoding="utf-8")
    return path


def limit_alternatives(data: dict, max_alternatives: int) -> dict:
    if max_alternatives <= 0 or len(data["alternatives"]) <= max_alternatives:
        return data
    kept = data["alternatives"][:max_alternatives]
    limited = copy.deepcopy(data)
    limited["alternatives"] = kept
    for dm in limited["decision_makers"]:
        limited["ratings"][dm] = {a: limited["ratings"][dm][a] for a in kept}
    return limited


def convert_supplier_workbook(path: Path, name: str) -> dict:
    """Convert the hesitant fuzzy TOPSIS supplier Excel examples."""
    ws = load_workbook(path, data_only=True)["Julgamentos DMs"]
    ratings: dict[str, dict] = {}
    alternatives: list[str] = []
    criteria: list[str] = []

    row = 1
    while row <= ws.max_row:
        dm = ws.cell(row, 8).value
        if isinstance(dm, str) and dm.startswith("DM"):
            dm_id = dm
            block_criteria = [
                str(ws.cell(row, col).value).replace("Neg ", "")
                for col in range(9, 13)
                if ws.cell(row, col).value
            ]
            if not criteria:
                criteria = block_criteria
            ratings[dm_id] = {}
            rr = row + 1
            while rr <= ws.max_row and ws.cell(rr, 8).value:
                alt = str(ws.cell(rr, 8).value)
                if alt.startswith("A"):
                    if alt not in alternatives:
                        alternatives.append(alt)
                    ratings[dm_id][alt] = {}
                    for idx, crit in enumerate(block_criteria, start=9):
                        parsed = parse_s_value(ws.cell(rr, idx).value)
                        if parsed is not None:
                            ratings[dm_id][alt][crit] = parsed
                rr += 1
            row = rr
        row += 1

    weights = {c: [1.0, 1.0, 1.0] for c in criteria}
    return {
        "source_name": name,
        "alternatives": alternatives,
        "criteria": criteria,
        "decision_makers": sorted(ratings),
        "criteria_weights": weights,
        "ratings": ratings,
    }


def read_csv_rows(path: Path) -> list[dict]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def read_xlsx_rows(path: Path, sheet_name: str | None = None) -> list[dict]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    out = []
    for row in rows[1:]:
        record = {}
        for h, v in zip(headers, row):
            if h:
                record[h] = v
        out.append(record)
    return out


def numeric_columns(rows: list[dict], exclude: set[str] | None = None) -> list[str]:
    exclude = {x.lower() for x in (exclude or set())}
    cols = []
    for key in rows[0]:
        if key.lower() in exclude:
            continue
        numeric_count = 0
        values = []
        for row in rows:
            try:
                values.append(float(row[key]))
                numeric_count += 1
            except (TypeError, ValueError):
                pass
        if numeric_count >= max(3, int(0.75 * len(rows))) and len(set(values)) > 1:
            cols.append(key)
    return cols


def convert_crisp_rows(
    rows: list[dict],
    name: str,
    alt_col: str,
    criteria: list[str],
    cost_criteria: set[str] | None = None,
    num_dms: int = 15,
    seed: int = 123,
) -> dict:
    """Convert crisp criteria to a synthetic fuzzy group-DM panel."""
    cost_criteria = cost_criteria or set()
    rng = random.Random(seed)
    alternatives = [str(row[alt_col]) for row in rows]
    values_by_crit = {}
    for crit in criteria:
        vals = []
        for row in rows:
            try:
                vals.append(float(row[crit]))
            except (TypeError, ValueError):
                vals.append(float("nan"))
        finite = [v for v in vals if math.isfinite(v)]
        lo, hi = min(finite), max(finite)
        values_by_crit[crit] = (vals, lo, hi)

    dms = [f"DM{i}" for i in range(1, num_dms + 1)]
    ratings = {dm: {} for dm in dms}
    for dm in dms:
        for alt_idx, alt in enumerate(alternatives):
            ratings[dm][alt] = {}
            for crit in criteria:
                vals, lo, hi = values_by_crit[crit]
                value = vals[alt_idx]
                if not math.isfinite(value) or abs(hi - lo) < 1e-12:
                    normalized = 0.5
                elif crit in cost_criteria:
                    normalized = (hi - value) / (hi - lo)
                else:
                    normalized = (value - lo) / (hi - lo)
                base_score = 1.0 + 8.0 * normalized
                noisy_score = base_score + rng.uniform(-0.45, 0.45)
                ratings[dm][alt][crit] = tfn_from_score(noisy_score)

    return {
        "source_name": name,
        "alternatives": alternatives,
        "criteria": criteria,
        "decision_makers": dms,
        "criteria_weights": {c: [1.0, 1.0, 1.0] for c in criteria},
        "ratings": ratings,
    }


def convert_car_dataset(path: Path, max_alternatives: int, seed: int) -> dict:
    cols = ["buying", "maint", "doors", "persons", "lug_boot", "safety", "class"]
    maps = {
        "buying": {"vhigh": 1, "high": 3, "med": 6, "low": 9},
        "maint": {"vhigh": 1, "high": 3, "med": 6, "low": 9},
        "doors": {"2": 1, "3": 4, "4": 7, "5more": 9},
        "persons": {"2": 1, "4": 5, "more": 9},
        "lug_boot": {"small": 1, "med": 5, "big": 9},
        "safety": {"low": 1, "med": 5, "high": 9},
    }
    rows = []
    with path.open(newline="", encoding="utf-8") as f:
        for idx, values in enumerate(csv.reader(f), start=1):
            row = dict(zip(cols, values))
            row["Alternative"] = f"Car_{idx}_{row['class']}"
            for c, mapping in maps.items():
                row[c] = mapping[row[c].replace("-", "")]
            rows.append(row)
    if max_alternatives > 0:
        rows = rows[:max_alternatives]
    return convert_crisp_rows(rows, "car_evaluation", "Alternative", list(maps), num_dms=15, seed=seed)


def load_real_datasets(max_alternatives: int, seed: int, include: set[str] | None = None) -> list[tuple[str, dict]]:
    datasets = []
    include = include or set()

    def wants(name: str) -> bool:
        return not include or name in include

    if wants("supplier_beg_rashid"):
        datasets.append((
            "supplier_beg_rashid",
            convert_supplier_workbook(REAL_DIR / "HesitantFuzzyTOPSIS Based Beg&Rashid.xlsx", "supplier_beg_rashid"),
        ))
    if wants("supplier_wu"):
        datasets.append((
            "supplier_wu",
            convert_supplier_workbook(REAL_DIR / "HesitantFuzzyTOPSIS Based Wu et al.xlsx", "supplier_wu"),
        ))

    if wants("healthcare_countries_2021"):
        country_rows = read_csv_rows(
            REAL_DIR / "Dataset for countries assessment regarding healthc" / "data" / "data_2021.csv"
        )
        country_criteria = [f"C{i}" for i in range(1, 14)]
        country_cost = {"C7", "C8", "C11", "C12", "C13"}
        datasets.append((
            "healthcare_countries_2021",
            convert_crisp_rows(
                country_rows,
                "healthcare_countries_2021",
                "Country",
                country_criteria,
                cost_criteria=country_cost,
                num_dms=15,
                seed=seed + 11,
            ),
        ))

    if wants("car_evaluation"):
        datasets.append((
            "car_evaluation",
            convert_car_dataset(REAL_DIR / "car+evaluation" / "car.data", max_alternatives, seed + 22),
        ))

    if wants("healthcare_resource_allocation"):
        hra_rows = read_xlsx_rows(REAL_DIR / "Healthcare Resource Allocation.xlsx", "基础回归")
        hra_criteria = numeric_columns(hra_rows, exclude={"code", "year"})
        for row in hra_rows:
            row["Alternative"] = f"{row.get('city', 'city')}_{row.get('year', '')}"
        datasets.append((
            "healthcare_resource_allocation",
            convert_crisp_rows(
                hra_rows[:max_alternatives] if max_alternatives > 0 else hra_rows,
                "healthcare_resource_allocation",
                "Alternative",
                hra_criteria,
                num_dms=15,
                seed=seed + 33,
            ),
        ))

    return [(name, limit_alternatives(data, max_alternatives)) for name, data in datasets]


def inject_synthetic_bias(data: dict, target_alt: str, attacker_fraction: float) -> tuple[dict, list[str]]:
    biased = copy.deepcopy(data)
    dms = biased["decision_makers"]
    num_attackers = max(1, round(len(dms) * attacker_fraction))
    attackers = dms[-num_attackers:]
    for dm in attackers:
        for alt in biased["alternatives"]:
            for crit in biased["criteria"]:
                biased["ratings"][dm][alt][crit] = [7.0, 9.0, 9.0] if alt == target_alt else [1.0, 1.0, 3.0]
    return biased, attackers


def ranking_order(result) -> list[str]:
    return [alt for alt, _ in result]


def rank_map(order: list[str]) -> dict[str, int]:
    return {alt: idx + 1 for idx, alt in enumerate(order)}


def spearman_rho(reference: list[str], candidate: list[str]) -> float:
    ref = rank_map(reference)
    cand = rank_map(candidate)
    n = len(reference)
    if n < 2:
        return 1.0
    d2 = sum((ref[a] - cand[a]) ** 2 for a in reference)
    return 1.0 - (6.0 * d2) / (n * (n * n - 1))


def kendall_tau(reference: list[str], candidate: list[str]) -> float:
    ref = rank_map(reference)
    sequence = [ref[a] for a in candidate]
    n = len(sequence)
    total_pairs = n * (n - 1) // 2
    if total_pairs == 0:
        return 1.0

    # Kendall tau for two complete rankings: discordant pairs are exactly the
    # inversions in the candidate sequence expressed in reference-rank space.
    bit = [0] * (n + 2)

    def bit_add(idx: int) -> None:
        while idx <= n:
            bit[idx] += 1
            idx += idx & -idx

    def bit_sum(idx: int) -> int:
        total = 0
        while idx > 0:
            total += bit[idx]
            idx -= idx & -idx
        return total

    inversions = 0
    seen = 0
    for value in sequence:
        inversions += seen - bit_sum(value)
        bit_add(value)
        seen += 1

    return 1.0 - (2.0 * inversions / total_pairs)


def summarize(values: list[float]) -> str:
    if not values:
        return "NA"
    if len(values) == 1:
        return f"{values[0]:.3f}"
    return f"{mean(values):.3f}±{pstdev(values):.3f}"


def evaluate_dataset(
    path: Path,
    reference_order: list[str],
    target_alt: str,
    args,
    dataset_name: str,
    mode: str,
) -> list[dict]:
    rows = []
    target_ref_rank = rank_map(reference_order)[target_alt]
    for method_name, runner, stochastic in METHODS:
        repeats = args.repeats if stochastic else 1
        print(
            f"  {dataset_name} [{mode}] {method_name}: "
            f"{repeats} run(s), {args.num_bags} bag(s)",
            flush=True,
        )
        metrics = []
        ranks = []
        top1 = []
        for repeat in range(repeats):
            seed = args.seed + repeat * 1000
            result = runner(str(path), args.num_bags, seed)
            order = ranking_order(result)
            ranks.append(rank_map(order)[target_alt])
            top1.append(1.0 if order[0] == reference_order[0] else 0.0)
            metrics.append({
                "spearman": spearman_rho(reference_order, order),
                "kendall": kendall_tau(reference_order, order),
                "target_rank_error": abs(rank_map(order)[target_alt] - target_ref_rank),
            })
        rows.append({
            "method": method_name,
            "spearman": summarize([m["spearman"] for m in metrics]),
            "kendall": summarize([m["kendall"] for m in metrics]),
            "top1_acc": summarize(top1),
            "target_rank": Counter(ranks).most_common(1)[0][0],
            "target_rank_error": summarize([m["target_rank_error"] for m in metrics]),
        })
    return rows


def write_csv(path: Path, rows: list[dict]) -> None:
    if not rows:
        return
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def print_metric_table(mode: str, rows: list[dict]) -> None:
    print(f"\n[{mode.upper()}] metrics vs clean M1 reference")
    print(f"{'Method':<8}{'Spearman':<16}{'Kendall':<16}{'Top1':<12}{'TargetRank':<12}{'TargetErr'}")
    for row in rows:
        print(
            f"{row['method']:<8}{row['spearman']:<16}{row['kendall']:<16}"
            f"{row['top1_acc']:<12}{row['target_rank']:<12}{row['target_rank_error']}"
        )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--num-bags", type=int, default=100)
    parser.add_argument("--repeats", type=int, default=10)
    parser.add_argument("--seed", type=int, default=20260425)
    parser.add_argument("--attacker-fraction", type=float, default=0.30)
    parser.add_argument("--max-alternatives", type=int, default=80,
                        help="Use 0 to run every alternative/row in large datasets.")
    parser.add_argument(
        "--datasets",
        nargs="*",
        default=None,
        help=(
            "Optional dataset names to run: supplier_beg_rashid, supplier_wu, "
            "healthcare_countries_2021, car_evaluation, healthcare_resource_allocation."
        ),
    )
    args = parser.parse_args()

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    all_summary = []
    selected = set(args.datasets or [])
    valid_datasets = {
        "supplier_beg_rashid",
        "supplier_wu",
        "healthcare_countries_2021",
        "car_evaluation",
        "healthcare_resource_allocation",
    }
    unknown = sorted(selected - valid_datasets)
    if unknown:
        raise SystemExit(f"Unknown dataset name(s): {', '.join(unknown)}")

    for dataset_name, clean_data in load_real_datasets(args.max_alternatives, args.seed, selected):
        clean_path = save_dataset(clean_data, f"{dataset_name}_clean")
        clean_reference = ranking_order(m1.run_m1(str(clean_path)))
        target_alt = clean_reference[-1]

        contaminated_data, attackers = inject_synthetic_bias(clean_data, target_alt, args.attacker_fraction)
        contaminated_path = save_dataset(contaminated_data, f"{dataset_name}_contaminated")

        print("\n" + "=" * 120)
        print(f"DATASET: {dataset_name}")
        print(f"Alternatives={len(clean_data['alternatives'])}, Criteria={len(clean_data['criteria'])}, DMs={len(clean_data['decision_makers'])}")
        print(f"Synthetic attack target: {target_alt} (clean M1 rank #{len(clean_reference)})")
        effective_attack = len(attackers) / len(clean_data["decision_makers"])
        print(f"Attackers: {attackers} ({effective_attack:.1%} effective contamination)")
        if len(clean_data["decision_makers"]) < 5:
            print("Note: very small DM panel; robustness results are diagnostic, not publication-grade.")

        clean_rows = evaluate_dataset(clean_path, clean_reference, target_alt, args, dataset_name, "clean")
        write_csv(OUT_DIR / f"{dataset_name}_clean_metrics.csv", clean_rows)
        print_metric_table("clean", clean_rows)
        for row in clean_rows:
            all_summary.append({"dataset": dataset_name, "mode": "clean", **row})

        contaminated_rows = evaluate_dataset(
            contaminated_path,
            clean_reference,
            target_alt,
            args,
            dataset_name,
            "contaminated",
        )
        write_csv(OUT_DIR / f"{dataset_name}_contaminated_metrics.csv", contaminated_rows)
        print_metric_table("contaminated", contaminated_rows)
        for row in contaminated_rows:
            all_summary.append({"dataset": dataset_name, "mode": "contaminated", **row})

    write_csv(OUT_DIR / "all_real_dataset_metrics.csv", all_summary)
    print("\n" + "=" * 120)
    print(f"Saved converted JSON datasets and metric CSVs to: {OUT_DIR}")


if __name__ == "__main__":
    main()
