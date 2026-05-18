from __future__ import annotations

import csv
import json
import math
import re
import tempfile
import time
from pathlib import Path
from typing import Any

import openpyxl
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles

import sys

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import m4_weighted_bagging as m4
import m6_reliability_weighted as m6
import m7_entropy_reliability as m7


app = FastAPI(
    title="EARR-TOPSIS API",
    description="Public API for running the three proposed reliability-aware fuzzy TOPSIS methods.",
    version="0.2.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["*"],
)

app.mount("/static", StaticFiles(directory=Path(__file__).parent / "static"), name="static")


METHODS = {
    "M4": {
        "name": "Reliability-Weighted Bagging",
        "purpose": "Samples reliable decision makers more often inside true bootstrap bags.",
    },
    "M6": {
        "name": "Reliability-Weighted Fuzzy Aggregation",
        "purpose": "Weights all TFN components by decision-maker reliability inside each bag.",
    },
    "M7": {
        "name": "EARR-TOPSIS",
        "purpose": "Combines entropy, variance-consistency, and clone/agreement reliability with weighted sampling, aggregation, and bag weighting.",
    },
}

HEADER_ALIASES = {
    "decision_maker": {"decision_maker", "decision maker", "dm", "expert", "judge", "evaluator", "respondent"},
    "alternative": {"alternative", "alt", "option", "item", "candidate", "supplier", "country", "name"},
    "criterion": {"criterion", "criteria", "attribute", "factor", "indicator", "metric"},
    "l": {"l", "lower", "low", "left", "min"},
    "m": {"m", "modal", "middle", "mean", "center", "centre"},
    "u": {"u", "upper", "high", "right", "max"},
    "weight_l": {"weight_l", "wl", "w_l", "lower_weight"},
    "weight_m": {"weight_m", "wm", "w_m", "modal_weight", "middle_weight"},
    "weight_u": {"weight_u", "wu", "w_u", "upper_weight"},
    "criteria_type": {"criteria_type", "criterion_type", "type", "benefit_cost", "benefit_or_cost"},
}

COST_KEYWORDS = {
    "cost",
    "price",
    "expense",
    "risk",
    "time",
    "delay",
    "latency",
    "distance",
    "loss",
    "mortality",
    "emission",
    "pollution",
}

DESCRIPTOR_COLUMNS = {
    "code",
    "id",
    "index",
    "year",
    "symbol",
    "region",
    "regin",
    "source",
    "adm",
    "oca",
    "sou",
}


def _clean_name(value: Any) -> str:
    text = str(value).strip()
    if not text:
        raise ValueError("Empty name found in dataset")
    return text


def _normalize_header(value: Any) -> str:
    text = re.sub(r"[^a-zA-Z0-9]+", "_", str(value).strip().lower()).strip("_")
    human = text.replace("_", " ")
    for canonical, aliases in HEADER_ALIASES.items():
        if text in aliases or human in aliases:
            return canonical
    return text


def _canonicalize_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    for row in rows:
        converted = {}
        for key, value in row.items():
            converted[_normalize_header(key)] = value
        out.append(converted)
    return out


def _as_float(value: Any, field: str) -> float:
    if value is None or value == "":
        raise ValueError(f"Missing numeric value for {field}")
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid numeric value for {field}: {value!r}") from exc
    if not math.isfinite(number):
        raise ValueError(f"Non-finite numeric value for {field}: {value!r}")
    return number


def _as_tfn(value: Any, field: str) -> list[float]:
    if isinstance(value, str):
        parts = [p for p in re.split(r"[,;| ]+", value.strip("()[] ")) if p]
        value = parts
    if not isinstance(value, (list, tuple)) or len(value) != 3:
        raise ValueError(f"{field} must be a triangular fuzzy number [l, m, u]")
    tfn = [_as_float(v, field) for v in value]
    if tfn[0] > tfn[1] or tfn[1] > tfn[2]:
        raise ValueError(f"{field} must satisfy l <= m <= u")
    return tfn


def _maybe_float(value: Any) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _read_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    header_index = 0
    for idx, row in enumerate(rows[:25]):
        non_empty = [cell for cell in row if cell not in (None, "")]
        if len(non_empty) >= 2:
            header_index = idx
            break
    rows = rows[header_index:]
    headers = [_normalize_header(v) for v in rows[0]]
    return _canonicalize_rows([
        {headers[i]: row[i] for i in range(len(headers))}
        for row in rows[1:]
        if any(cell is not None for cell in row)
    ])


def _read_xls_rows(path: Path) -> list[dict[str, Any]]:
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError("Legacy .xls upload requires xlrd. Install web_api/requirements.txt.") from exc
    workbook = xlrd.open_workbook(str(path))
    sheet = workbook.sheet_by_index(0)
    rows = []
    for r in range(sheet.nrows):
        rows.append(sheet.row_values(r))
    if not rows:
        return []
    header_index = 0
    for idx, row in enumerate(rows[:25]):
        non_empty = [cell for cell in row if cell not in (None, "")]
        if len(non_empty) >= 2:
            header_index = idx
            break
    rows = rows[header_index:]
    headers = [_normalize_header(v) for v in rows[0]]
    return _canonicalize_rows([
        {headers[i]: row[i] for i in range(len(headers))}
        for row in rows[1:]
        if any(cell not in (None, "") for cell in row)
    ])


def _read_delimited_rows(path: Path) -> list[dict[str, Any]]:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    with path.open(newline="", encoding="utf-8-sig") as f:
        return _canonicalize_rows(list(csv.DictReader(f, delimiter=delimiter)))


def _ordered_unique(values: list[str]) -> list[str]:
    seen = set()
    result = []
    for value in values:
        if value not in seen:
            seen.add(value)
            result.append(value)
    return result


def _flat_rows_to_dataset(rows: list[dict[str, Any]]) -> dict:
    required = {"decision_maker", "alternative", "criterion", "l", "m", "u"}
    if not rows or not required.issubset(rows[0]):
        raise ValueError("Flat fuzzy table must contain decision_maker, alternative, criterion, l, m, u")

    decision_makers = _ordered_unique([_clean_name(r["decision_maker"]) for r in rows])
    alternatives = _ordered_unique([_clean_name(r["alternative"]) for r in rows])
    criteria = _ordered_unique([_clean_name(r["criterion"]) for r in rows])
    ratings = {
        dm: {alt: {crit: [0.0, 0.0, 0.0] for crit in criteria} for alt in alternatives}
        for dm in decision_makers
    }
    criteria_weights = {crit: [1.0, 1.0, 1.0] for crit in criteria}
    criteria_types = {}

    for row in rows:
        dm = _clean_name(row["decision_maker"])
        alt = _clean_name(row["alternative"])
        crit = _clean_name(row["criterion"])
        ratings[dm][alt][crit] = [
            _as_float(row["l"], f"{dm}/{alt}/{crit}/l"),
            _as_float(row["m"], f"{dm}/{alt}/{crit}/m"),
            _as_float(row["u"], f"{dm}/{alt}/{crit}/u"),
        ]
        if row.get("weight_l") not in (None, ""):
            criteria_weights[crit] = [
                _as_float(row["weight_l"], f"{crit}/weight_l"),
                _as_float(row.get("weight_m", row["weight_l"]), f"{crit}/weight_m"),
                _as_float(row.get("weight_u", row.get("weight_m", row["weight_l"])), f"{crit}/weight_u"),
            ]
        if row.get("criteria_type") not in (None, ""):
            criteria_types[crit] = str(row["criteria_type"]).lower()

    data = {
        "alternatives": alternatives,
        "criteria": criteria,
        "decision_makers": decision_makers,
        "criteria_weights": criteria_weights,
        "ratings": ratings,
    }
    if criteria_types:
        data["criteria_types"] = criteria_types
    return data


def _infer_cost_criteria(criteria: list[str], requested: set[str]) -> set[str]:
    lowered_requested = {c.strip().lower() for c in requested}
    inferred = set()
    for crit in criteria:
        low = crit.lower()
        tokens = set(re.split(r"[^a-z0-9]+", low))
        if low in lowered_requested or tokens & COST_KEYWORDS:
            inferred.add(crit)
    return inferred


def _row_label(row: dict[str, Any], alt_col: str) -> str:
    if "city" in row and "year" in row and row.get("city") not in (None, "") and row.get("year") not in (None, ""):
        return f"{_clean_name(row['city'])}_{_clean_name(row['year'])}"
    if "country" in row and row.get("country") not in (None, ""):
        return _clean_name(row["country"])
    return _clean_name(row[alt_col])


def _crisp_rows_to_dataset(
    rows: list[dict[str, Any]],
    pseudo_dms: int = 15,
    cost_criteria: set[str] | None = None,
) -> dict:
    if not rows:
        raise ValueError("Empty crisp table")
    headers = list(rows[0])
    alt_col = "alternative" if "alternative" in headers else None
    if alt_col is None:
        for header in headers:
            non_numeric = sum(_maybe_float(row.get(header)) is None for row in rows)
            if non_numeric >= max(1, len(rows) // 2):
                alt_col = header
                break
    alt_col = alt_col or headers[0]
    criteria = [
        h for h in headers
        if h != alt_col
        and h not in DESCRIPTOR_COLUMNS
        and sum(_maybe_float(row.get(h)) is not None for row in rows) >= max(2, int(0.75 * len(rows)))
        and len({row.get(h) for row in rows if _maybe_float(row.get(h)) is not None}) > 1
    ]
    if not criteria:
        raise ValueError("Crisp table must contain at least one numeric criterion column")
    if pseudo_dms < 1 or pseudo_dms > 200:
        raise ValueError("pseudo_dms must be between 1 and 200")

    cost_set = _infer_cost_criteria(criteria, cost_criteria or set())
    alternatives = [_row_label(row, alt_col) for row in rows]
    values = {
        crit: [_maybe_float(row.get(crit)) for row in rows]
        for crit in criteria
    }
    scaled = {}
    for crit, vals in values.items():
        finite_vals = [v for v in vals if v is not None]
        if not finite_vals:
            continue
        lo, hi = min(finite_vals), max(finite_vals)
        if abs(hi - lo) < 1e-12:
            scaled[crit] = [5.0 for _ in vals]
        elif crit in cost_set:
            scaled[crit] = [5.0 if v is None else 1.0 + 8.0 * (hi - v) / (hi - lo) for v in vals]
        else:
            scaled[crit] = [5.0 if v is None else 1.0 + 8.0 * (v - lo) / (hi - lo) for v in vals]

    decision_makers = [f"DM{i + 1}" for i in range(pseudo_dms)]
    ratings = {
        dm: {alt: {crit: [0.0, 0.0, 0.0] for crit in criteria} for alt in alternatives}
        for dm in decision_makers
    }
    for dm_index, dm in enumerate(decision_makers):
        phase = (dm_index + 1) / pseudo_dms
        for alt_index, alt in enumerate(alternatives):
            for crit in criteria:
                jitter = 0.25 * math.sin((alt_index + 1) * (dm_index + 1))
                mid = min(9.0, max(1.0, scaled[crit][alt_index] + jitter * phase))
                ratings[dm][alt][crit] = [max(1.0, mid - 0.5), mid, min(9.0, mid + 0.5)]

    return {
        "alternatives": alternatives,
        "criteria": criteria,
        "decision_makers": decision_makers,
        "criteria_weights": {crit: [1.0, 1.0, 1.0] for crit in criteria},
        "ratings": ratings,
        "conversion": {
            "source": "crisp_matrix",
            "pseudo_decision_makers": pseudo_dms,
            "cost_criteria_scaled_as_benefit": sorted(cost_set),
        },
    }


def _parse_s_value(value: Any) -> list[float] | None:
    if not isinstance(value, str) or not value.strip().startswith("[s"):
        return None
    nums = []
    for part in value.strip("[]").split(","):
        part = part.strip()
        if part.startswith("s"):
            nums.append(float(part[1:]))
    if not nums:
        return None
    lo, hi = min(nums), max(nums)
    mid = sum(nums) / len(nums)
    def scale(x: float) -> float:
        return 1.0 + (8.0 * x / 6.0)
    return [round(scale(lo), 4), round(scale(mid), 4), round(scale(hi), 4)]


def _supplier_workbook_to_dataset(path: Path) -> dict:
    workbook = openpyxl.load_workbook(path, data_only=True)
    if "Julgamentos DMs" not in workbook.sheetnames:
        raise ValueError("Not a recognized hesitant fuzzy supplier workbook")
    ws = workbook["Julgamentos DMs"]
    ratings: dict[str, dict[str, dict[str, list[float]]]] = {}
    alternatives: list[str] = []
    criteria: list[str] = []

    row = 1
    while row <= ws.max_row:
        dm = ws.cell(row, 8).value
        if isinstance(dm, str) and dm.startswith("DM"):
            dm_id = dm.strip()
            block_criteria = [
                str(ws.cell(row, col).value).replace("Neg ", "").strip()
                for col in range(9, 13)
                if ws.cell(row, col).value
            ]
            if not criteria:
                criteria = block_criteria
            ratings[dm_id] = {}
            rr = row + 1
            while rr <= ws.max_row and ws.cell(rr, 8).value:
                alt = str(ws.cell(rr, 8).value).strip()
                if alt.startswith("A"):
                    if alt not in alternatives:
                        alternatives.append(alt)
                    ratings[dm_id][alt] = {}
                    for idx, crit in enumerate(block_criteria, start=9):
                        parsed = _parse_s_value(ws.cell(rr, idx).value)
                        if parsed is not None:
                            ratings[dm_id][alt][crit] = parsed
                rr += 1
            row = rr
        row += 1

    if not alternatives or not criteria or not ratings:
        raise ValueError("Could not extract decision-maker blocks from hesitant fuzzy workbook")

    return {
        "alternatives": alternatives,
        "criteria": criteria,
        "decision_makers": sorted(ratings),
        "criteria_weights": {crit: [1.0, 1.0, 1.0] for crit in criteria},
        "ratings": ratings,
        "conversion": {
            "source": "hesitant_fuzzy_supplier_workbook",
            "sheet": "Julgamentos DMs",
        },
    }


def _validate_native_dataset(data: dict) -> dict:
    if not isinstance(data, dict):
        raise ValueError("Native JSON dataset must be an object")
    for key in ["alternatives", "criteria", "decision_makers", "ratings"]:
        if key not in data:
            raise ValueError(f"Missing required key: {key}")

    alternatives = [_clean_name(a) for a in data["alternatives"]]
    criteria = [_clean_name(c) for c in data["criteria"]]
    decision_makers = [_clean_name(dm) for dm in data["decision_makers"]]
    if not alternatives or not criteria or not decision_makers:
        raise ValueError("Dataset must contain at least one alternative, criterion, and decision maker")

    criteria_weights = data.get("criteria_weights") or {crit: [1.0, 1.0, 1.0] for crit in criteria}
    clean_weights = {crit: _as_tfn(criteria_weights.get(crit, [1.0, 1.0, 1.0]), f"criteria_weights/{crit}") for crit in criteria}

    ratings = data["ratings"]
    clean_ratings = {}
    for dm in decision_makers:
        if dm not in ratings:
            raise ValueError(f"Missing ratings for decision maker {dm}")
        clean_ratings[dm] = {}
        for alt in alternatives:
            if alt not in ratings[dm]:
                raise ValueError(f"Missing ratings for {dm}/{alt}")
            clean_ratings[dm][alt] = {}
            for crit in criteria:
                if crit not in ratings[dm][alt]:
                    raise ValueError(f"Missing rating for {dm}/{alt}/{crit}")
                clean_ratings[dm][alt][crit] = _as_tfn(ratings[dm][alt][crit], f"{dm}/{alt}/{crit}")

    clean = {
        "alternatives": alternatives,
        "criteria": criteria,
        "decision_makers": decision_makers,
        "criteria_weights": clean_weights,
        "ratings": clean_ratings,
    }
    if isinstance(data.get("criteria_types"), dict):
        clean["criteria_types"] = {str(k): str(v).lower() for k, v in data["criteria_types"].items()}
    if isinstance(data.get("conversion"), dict):
        clean["conversion"] = data["conversion"]
    return clean


def _limit_alternatives(data: dict, max_alternatives: int) -> dict:
    if max_alternatives <= 0 or len(data["alternatives"]) <= max_alternatives:
        return data
    kept = data["alternatives"][:max_alternatives]
    kept_set = set(kept)
    limited = dict(data)
    limited["alternatives"] = kept
    limited["ratings"] = {
        dm: {alt: ratings[alt] for alt in kept if alt in ratings}
        for dm, ratings in data["ratings"].items()
    }
    conversion = dict(data.get("conversion") or {})
    conversion["limited_from_alternatives"] = len(data["alternatives"])
    conversion["limited_to_alternatives"] = len(kept_set)
    limited["conversion"] = conversion
    return limited


def parse_upload(path: Path, pseudo_dms: int = 15, cost_criteria: set[str] | None = None) -> dict:
    suffix = path.suffix.lower()
    if suffix == ".json":
        data = _validate_native_dataset(json.loads(path.read_text(encoding="utf-8")))
    elif suffix in {".csv", ".tsv"}:
        rows = _read_delimited_rows(path)
        if not rows:
            raise ValueError("Uploaded table is empty")
        data = _flat_rows_to_dataset(rows) if "decision_maker" in rows[0] else _crisp_rows_to_dataset(rows, pseudo_dms, cost_criteria)
    elif suffix in {".xlsx", ".xlsm"}:
        try:
            data = _supplier_workbook_to_dataset(path)
        except ValueError:
            rows = _read_xlsx_rows(path)
            if not rows:
                raise ValueError("Uploaded workbook is empty")
            data = _flat_rows_to_dataset(rows) if "decision_maker" in rows[0] else _crisp_rows_to_dataset(rows, pseudo_dms, cost_criteria)
    elif suffix == ".xls":
        rows = _read_xls_rows(path)
        if not rows:
            raise ValueError("Uploaded workbook is empty")
        data = _flat_rows_to_dataset(rows) if "decision_maker" in rows[0] else _crisp_rows_to_dataset(rows, pseudo_dms, cost_criteria)
    else:
        raise ValueError("Supported formats: JSON, CSV, TSV, XLS, XLSX, XLSM")

    return _validate_native_dataset(data)


def _json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(k): _json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_json_safe(v) for v in value]
    if hasattr(value, "item"):
        return value.item()
    if isinstance(value, float):
        return round(value, 10)
    return value


def run_method(method: str, dataset_path: Path, num_bags: int, seed: int, top_k: int) -> dict:
    start = time.perf_counter()
    if method == "M4":
        ranked, reliability = m4.run_m4(str(dataset_path), num_bags=num_bags, seed=seed)
    elif method == "M6":
        ranked, reliability = m6.run_m6(str(dataset_path), num_bags=num_bags, seed=seed)
    elif method == "M7":
        ranked, reliability = m7.run_m7(str(dataset_path), num_bags=num_bags, seed=seed)
    else:
        raise ValueError(f"Unknown method: {method}")
    elapsed = time.perf_counter() - start

    ranked_rows = [
        {"rank": i + 1, "alternative": alt, "score": float(score)}
        for i, (alt, score) in enumerate(ranked)
    ]
    reliability_rows = [
        {"decision_maker": dm, "reliability": float(score)}
        for dm, score in sorted(reliability.items(), key=lambda item: item[1], reverse=True)
    ]

    return {
        "method": method,
        "name": METHODS[method]["name"],
        "runtime_seconds": round(elapsed, 6),
        "top_alternative": ranked_rows[0]["alternative"] if ranked_rows else None,
        "ranking": ranked_rows[:top_k],
        "ranking_count": len(ranked_rows),
        "reliability": reliability_rows,
        "reliability_count": len(reliability_rows),
    }


@app.get("/", response_class=HTMLResponse)
def index() -> str:
    return (Path(__file__).parent / "static" / "index.html").read_text(encoding="utf-8")


@app.get("/health")
def health() -> dict:
    return {"status": "ok", "service": "EARR-TOPSIS API", "version": app.version}


@app.get("/api/methods")
def methods() -> dict:
    return {"methods": METHODS}


@app.get("/api/schema")
def schema() -> dict:
    return {
        "native_json": {
            "required": ["alternatives", "criteria", "decision_makers", "ratings"],
            "optional": ["criteria_weights", "criteria_types"],
        },
        "flat_fuzzy_table": {
            "required_columns": ["decision_maker", "alternative", "criterion", "l", "m", "u"],
            "optional_columns": ["weight_l", "weight_m", "weight_u", "criteria_type"],
        },
        "crisp_matrix": {
            "rule": "First text column is treated as alternatives; remaining numeric columns become criteria.",
            "options": ["pseudo_dms", "cost_criteria"],
        },
    }


@app.post("/api/run")
async def run_api(
    file: UploadFile = File(...),
    methods: str = Form("M4,M6,M7"),
    num_bags: int = Form(100),
    seed: int = Form(42),
    pseudo_dms: int = Form(15),
    cost_criteria: str = Form(""),
    top_k: int = Form(50),
    max_alternatives: int = Form(300),
) -> dict:
    selected = [m.strip().upper() for m in methods.split(",") if m.strip()]
    invalid = [m for m in selected if m not in METHODS]
    if invalid:
        raise HTTPException(status_code=400, detail=f"Unsupported methods: {invalid}")
    if num_bags < 1 or num_bags > 2000:
        raise HTTPException(status_code=400, detail="num_bags must be between 1 and 2000")
    if top_k < 1 or top_k > 1000:
        raise HTTPException(status_code=400, detail="top_k must be between 1 and 1000")

    suffix = Path(file.filename or "upload.json").suffix or ".json"
    cost_set = {c.strip() for c in cost_criteria.split(",") if c.strip()}
    with tempfile.TemporaryDirectory() as tmp:
        upload_path = Path(tmp) / f"dataset{suffix}"
        upload_path.write_bytes(await file.read())
        try:
            data = parse_upload(upload_path, pseudo_dms=pseudo_dms, cost_criteria=cost_set)
            data = _limit_alternatives(data, max_alternatives)
            dataset_path = Path(tmp) / "dataset.json"
            dataset_path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
            results = {
                method: run_method(method, dataset_path, num_bags=num_bags, seed=seed, top_k=top_k)
                for method in selected
            }
        except Exception as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    return _json_safe({
        "metadata": {
            "filename": file.filename,
            "alternatives": len(data["alternatives"]),
            "criteria": len(data["criteria"]),
            "decision_makers": len(data["decision_makers"]),
            "num_bags": num_bags,
            "seed": seed,
            "top_k": top_k,
            "max_alternatives": max_alternatives,
            "criteria_types": data.get("criteria_types", {}),
            "conversion": data.get("conversion"),
        },
        "results": results,
    })


@app.post("/api/validate")
async def validate_api(
    file: UploadFile = File(...),
    pseudo_dms: int = Form(15),
    cost_criteria: str = Form(""),
    max_alternatives: int = Form(300),
) -> dict:
    suffix = Path(file.filename or "upload.json").suffix or ".json"
    cost_set = {c.strip() for c in cost_criteria.split(",") if c.strip()}
    with tempfile.TemporaryDirectory() as tmp:
        upload_path = Path(tmp) / f"dataset{suffix}"
        upload_path.write_bytes(await file.read())
        try:
            data = parse_upload(upload_path, pseudo_dms=pseudo_dms, cost_criteria=cost_set)
            data = _limit_alternatives(data, max_alternatives)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    return _json_safe({
        "valid": True,
        "metadata": {
            "filename": file.filename,
            "alternatives": len(data["alternatives"]),
            "criteria_count": len(data["criteria"]),
            "decision_makers": len(data["decision_makers"]),
            "criteria": data["criteria"],
            "criteria_types": data.get("criteria_types", {}),
            "conversion": data.get("conversion"),
        },
    })


@app.get("/api/example")
def example() -> dict:
    return {
        "alternatives": ["A1", "A2"],
        "criteria": ["C1", "C2"],
        "decision_makers": ["DM1", "DM2", "DM3"],
        "criteria_weights": {"C1": [1, 1, 1], "C2": [1, 1, 1]},
        "ratings": {
            "DM1": {"A1": {"C1": [6, 7, 8], "C2": [5, 6, 7]}, "A2": {"C1": [4, 5, 6], "C2": [5, 5, 6]}},
            "DM2": {"A1": {"C1": [6, 7, 9], "C2": [5, 7, 8]}, "A2": {"C1": [4, 5, 6], "C2": [4, 5, 6]}},
            "DM3": {"A1": {"C1": [5, 6, 7], "C2": [5, 6, 8]}, "A2": {"C1": [5, 5, 6], "C2": [4, 5, 7]}},
        },
    }
